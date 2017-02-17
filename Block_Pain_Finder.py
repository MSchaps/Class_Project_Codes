def Block_Pair_Finder(directory):
    import arcpy
    import openpyxl
    import string
    import os
    from openpyxl import load_workbook, Workbook
    for file in os.listdir(directory):
        if file.endswith(".xlsx"):
            pairnum = 0
            Home_Blocks = []
            Work_Blocks = []
            print(file)
            wb = load_workbook(directory + r'\\' + file, read_only=True)
            ws = wb.worksheets[0]
            maxrow = ws.max_row
            Alphabet = list(string.ascii_uppercase)
            SheetNameDict = {0:"TOT_JOBS", 1:"JOBS_29_UND", 2:"JOBS_30_54", 3:"JOBS_55_OLD", 4:"EARN_1250_UND", 5:"EARN_1251_3333", \
                             6:"EARN_3334_UP", 7:"GOODS_PROC", 8:"TRAD_TRAN_UT", 9:"ALL_OTH"}
            outputsheetnum = 0
            newwb = Workbook()
            for num in range(2,12):
                letter = Alphabet[num]
                Numbers = []
                for i in range(2,maxrow + 1):
                    target_counties = ("27003", "27019", "27037", "27053", "27123", "27139", "27163")
                    if any(ws.cell(row=i, column=1).value.startswith(tc) for tc in target_counties):
                        print "Matching Home Block Found With " + ws.cell(row=i, column=1).value
                        if any(ws.cell(row=i, column=2).value.startswith(tc) for tc in target_counties):
                            print "Matching Work Block Found with " + ws.cell(row=i, column=2).value
                            pairnum += 1
                            print pairnum
                            HomeBlock = ws['A{0}'. format(i)].value
                            Home_Blocks.append(HomeBlock)
                            WorkBlock= ws['B{0}'. format(i)].value
                            Work_Blocks.append(WorkBlock)
                            Worker_Count = ws['{0}{1}'.format(letter, i)]
                            Numbers.append(Worker_Count)
                Block_Pairs = zip(Home_Blocks, Work_Blocks)
                lenbp = len(Block_Pairs)
                Worker_num_pair = zip(Block_Pairs, Numbers)
                SheetName = SheetNameDict[outputsheetnum]
                newws = wb.create_sheet(SheetName, outputsheetnum)
                outputsheetnum += 1
                for bp in Block_Pairs:
                    numworkers = 0
                    uniquehbp = bp[0]
                    for n in range(2, lenbp + 1):
                        for row in Worker_num_pair:
                            hb = row[0][0]
                            total = row[1]
                            if uniquehbp == hb:
                                newws["A{0}".format(n)] = uniquehbp
                                numworkers += total
                                print "Total Number of workers for pair is " + str(total)
                    print numworkers
                    newws["B{0}".format(n)] = numworkers
                newwb.save(directory + r'\\' + file  + r'.xlsx')
directory = r'C:\\GIS\\Cyber_GIS\\Final_Project\\Data\\LODES_OD_2014_MAIN\\Finals'
Block_Pair_Finder(directory)